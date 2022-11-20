# Author: Dakota Funke

from openpyxl import load_workbook
from openpyxl.styles import Font
import math
import os

# Future Additions:
# * Macro Implementation to turn certain subsection of rows red if already chosen
# * Perhaps Make or other command functionality for program to assist in easier processing
# * Singlular Generation of Excel Document by File Specification rather than pulling all

CONST_RANK_NUM = {
    'Iron IV': 1,
    'Iron III': 2,
    'Iron II': 3,
    'Iron I': 4,
    'Bronze IV': 5,
    'Bronze III': 6,
    'Bronze II': 7,
    'Bronze I': 8,
    'Silver IV': 9,
    'Silver III': 10,
    'Silver II': 11,
    'Silver I': 12,
    'Gold IV': 13,
    'Gold III': 14,
    'Gold II': 15,
    'Gold I': 16,
    'Platinum IV': 17,
    'Platinum III': 18,
    'Platinum II': 19,
    'Platinum I': 20,
    'Diamond IV': 21,
    'Diamond III': 22,
    'Diamond II': 23,
    'Diamond I': 24,
    'Master': 25,
    'Grandmaster': 26,
    'Challenger': 27
}

def GetFileList():
    # Desc: Gets all .xlsx files within the InputForms folder. Returns those files and corresponding data.

    files = []
    cwd = os.path.abspath(os.getcwd())
    dataPath = cwd + "/InputForms"
    for results in os.walk(dataPath):
        for file in results[2]:
            if(".xlsx" in file):
               files.append([os.path.join(dataPath, file), file])
    return files

def LoadWorkbook(fullPath, fileName):
    # Desc: Pulls the data from the file on the filepath into a usable Openpyxl Workbook Object

    wb = load_workbook(fullPath)
    WorkbookInfo = (fullPath, fileName, wb)
    return WorkbookInfo

def GetValidEntries(worksheet, rowMax):
    # Desc: Checks for all valid users by whether or not there is an x within the 'I' Column

    validData = []
    i = 1
    while(i != rowMax):
        if (worksheet.cell(row=i, column=9).value == 'x') or (worksheet.cell(row=i, column=9).value == 'X'):
            validData.append([worksheet.cell(row=i, column=2).value, 
                              worksheet.cell(row=i, column=3).value,
                              worksheet.cell(row=i, column=4).value,
                              worksheet.cell(row=i, column=5).value,
                              worksheet.cell(row=i, column=6).value,
                              worksheet.cell(row=i, column=7).value,
                              worksheet.cell(row=i, column=8).value])
        i += 1
    return validData

def CreateTeamWorksheet(userData, workbook):
    # Desc: Creates a worksheet that creates the maximum number of teams based off of the number of valid users.
    # Also includes the Discord Username and In-Game Names of the users willing to be Team Captains.

    teamSheet = workbook.create_sheet("Team Info")
    numTeams = 0
    teamIndex = 1
    currentRow = 1                                 # OpenPyxl Starts Reference @ 1
    currentCol = 1                                 # OpenPyxl Starts Reference @ 1
    currentRowAvailable = 3
    currentRowBackup = 3
    availableCaptainColumn = 11
    backupCaptainColumn = 15

    numTeams = math.ceil(len(userData) / 5)

    # Minimize User Data to Potential Team Captains
    captainUserData = [[0, 'None', 'None', 'None', 'None']]
    for user in userData:
        if user[4] != 'No':
            userNum = 0
            userNum = CONST_RANK_NUM[user[2]]
            if userNum != 0:
                if len(captainUserData) == 0:
                    captainUserData.append([userNum, user[2], user[0], user[1], user[4]])
                else:
                    dataIndex = 0
                    for index, captainUser in enumerate(captainUserData):
                        if captainUser[0] >= userNum:
                            dataIndex = index
                            break
                    captainUserData.insert(dataIndex+1, [userNum, user[2], user[0], user[1], user[4]])
            else:
                print("Unrecognized Rank '{rank}'. Check Data".format(rank = user[2]))

    # Create Main Team Information Areas
    while teamIndex <= numTeams:
        if teamIndex % 2 == 1: 
            currentRow = (math.floor(teamIndex / 2) * 8) + 1
            currentCol = 1
        else:
            currentCol = 6

        teamSheet.cell(row=currentRow, column=currentCol).font = Font(name="Roboto Mono", bold=True)
        teamSheet.cell(row=currentRow, column=currentCol).value = "Team "+str(teamIndex)
        teamSheet.cell(row=currentRow+1, column=currentCol).font = Font(name="Roboto Mono", underline="single")
        teamSheet.cell(row=currentRow+1, column=currentCol).value = "Lane"
        teamSheet.cell(row=currentRow+1, column=currentCol+1).font = Font(name="Roboto Mono", underline="single")
        teamSheet.cell(row=currentRow+1, column=currentCol+1).value = "Discord Name"
        teamSheet.cell(row=currentRow+1, column=currentCol+2).font = Font(name="Roboto Mono", underline="single")
        teamSheet.cell(row=currentRow+1, column=currentCol+2).value = "In-Game Name"
        teamSheet.cell(row=currentRow+1, column=currentCol+3).font = Font(name="Roboto Mono", underline="single")
        teamSheet.cell(row=currentRow+1, column=currentCol+3).value = "Rank"
        teamSheet.cell(row=currentRow+2, column=currentCol).font = Font(name="Roboto Mono")
        teamSheet.cell(row=currentRow+2, column=currentCol).value = "Top"
        teamSheet.cell(row=currentRow+3, column=currentCol).font = Font(name="Roboto Mono")
        teamSheet.cell(row=currentRow+3, column=currentCol).value = "Jungle"
        teamSheet.cell(row=currentRow+4, column=currentCol).font = Font(name="Roboto Mono")
        teamSheet.cell(row=currentRow+4, column=currentCol).value = "Mid"
        teamSheet.cell(row=currentRow+5, column=currentCol).font = Font(name="Roboto Mono")
        teamSheet.cell(row=currentRow+5, column=currentCol).value = "ADC"
        teamSheet.cell(row=currentRow+6, column=currentCol).font = Font(name="Roboto Mono")
        teamSheet.cell(row=currentRow+6, column=currentCol).value = "Sup"
        teamIndex += 1

    # Create Team Captain Information Areas
    teamSheet.cell(row=currentRowAvailable-2, column=availableCaptainColumn).font = Font(name="Roboto Mono", bold=True)
    teamSheet.cell(row=currentRowAvailable-2, column=availableCaptainColumn).value = "Available Team Captains"
    teamSheet.cell(row=currentRowAvailable-1, column=availableCaptainColumn).font = Font(name="Roboto Mono", underline="single")
    teamSheet.cell(row=currentRowAvailable-1, column=availableCaptainColumn).value = "Rank"
    teamSheet.cell(row=currentRowAvailable-1, column=availableCaptainColumn+1).font = Font(name="Roboto Mono", underline="single")
    teamSheet.cell(row=currentRowAvailable-1, column=availableCaptainColumn+1).value = "Discord Name"
    teamSheet.cell(row=currentRowAvailable-1, column=availableCaptainColumn+2).font = Font(name="Roboto Mono", underline="single")
    teamSheet.cell(row=currentRowAvailable-1, column=availableCaptainColumn+2).value = "Discord Name"

    teamSheet.cell(row=currentRowAvailable-2, column=backupCaptainColumn).font = Font(name="Roboto Mono", bold=True)
    teamSheet.cell(row=currentRowAvailable-2, column=backupCaptainColumn).value = "Backup Team Captains"
    teamSheet.cell(row=currentRowAvailable-1, column=backupCaptainColumn).font = Font(name="Roboto Mono", underline="single")
    teamSheet.cell(row=currentRowAvailable-1, column=backupCaptainColumn).value = "Rank"
    teamSheet.cell(row=currentRowAvailable-1, column=backupCaptainColumn+1).font = Font(name="Roboto Mono", underline="single")
    teamSheet.cell(row=currentRowAvailable-1, column=backupCaptainColumn+1).value = "Discord Name"
    teamSheet.cell(row=currentRowAvailable-1, column=backupCaptainColumn+2).font = Font(name="Roboto Mono", underline="single")
    teamSheet.cell(row=currentRowAvailable-1, column=backupCaptainColumn+2).value = "Discord Name"

    for captainData in captainUserData:
        if captainData[4] == 'Yes':
            teamSheet.cell(row=currentRowAvailable, column=availableCaptainColumn).value = captainData[1]
            teamSheet.cell(row=currentRowAvailable, column=availableCaptainColumn+1).value = captainData[2]
            teamSheet.cell(row=currentRowAvailable, column=availableCaptainColumn+2).value = captainData[3]
            currentRowAvailable += 1
        elif captainData[4] == 'Maybe':
            teamSheet.cell(row=currentRowBackup, column=backupCaptainColumn).value = captainData[1]
            teamSheet.cell(row=currentRowBackup, column=backupCaptainColumn+1).value = captainData[2]
            teamSheet.cell(row=currentRowBackup, column=backupCaptainColumn+2).value = captainData[3]
            currentRowBackup += 1

    # Modify Column Dimensions to Look More Visually Pleasing
    teamSheet.column_dimensions['B'].width = 15
    teamSheet.column_dimensions['C'].width = 15
    teamSheet.column_dimensions['D'].width = 20
    teamSheet.column_dimensions['G'].width = 15
    teamSheet.column_dimensions['H'].width = 15
    teamSheet.column_dimensions['I'].width = 20
    teamSheet.column_dimensions['K'].width = 15
    teamSheet.column_dimensions['L'].width = 25
    teamSheet.column_dimensions['M'].width = 25
    teamSheet.column_dimensions['O'].width = 15
    teamSheet.column_dimensions['P'].width = 25
    teamSheet.column_dimensions['Q'].width = 25

    print("Successfully Generated Team & Captain Info Sheet!")
    return

def CreatePrimaryRoleWorksheet(userData, workbook):
    # Desc: Creates a worksheet that lists all users by their selected Primary Roles and then Ranks
    
    primarySheet = workbook.create_sheet("Primary Roles")
    currentRow = 1
    currentCol = 1
    adcRow = 6
    supRow = 6
    fillRow = 6
    topPlayers = [[0, 'None', 'None', 'None', 'None']]
    junglePlayers = [[0, 'None', 'None', 'None', 'None']]
    midPlayers = [[0, 'None', 'None', 'None', 'None']]
    adcPlayers = [[0, 'None', 'None', 'None', 'None']]
    supPlayers = [[0, 'None', 'None', 'None', 'None']]
    fillPlayers = [[0, 'None', 'None', 'None', 'None']]

    # Sort Valid Players into their respective ranks
    for user in userData:
        userNum = 0
        userNum = CONST_RANK_NUM[user[2]]
        if userNum != 0:
            if user[5] == 'Top Lane': 
                if len(topPlayers) == 0:
                    topPlayers.append([userNum, user[0], user[1], user[2], user[3]]) # Discord Name, In-Game Name, Rank, LP
                else:
                    dataIndex = 0
                    for index, player in enumerate(topPlayers):
                        if player[0] < userNum:
                            dataIndex = index
                            break
                    topPlayers.insert(dataIndex, [userNum, user[0], user[1], user[2], user[3]])
            elif user[5] == 'Jungle': 
                if len(junglePlayers) == 0:
                    junglePlayers.append([userNum, user[0], user[1], user[2], user[3]]) # Discord Name, In-Game Name, Rank, LP
                else:
                    dataIndex = 0
                    for index, player in enumerate(junglePlayers):
                        if player[0] < userNum:
                            dataIndex = index
                            break
                    junglePlayers.insert(dataIndex, [userNum, user[0], user[1], user[2], user[3]])
            elif user[5] == 'Mid Lane': 
                if len(midPlayers) == 0:
                    midPlayers.append([userNum, user[0], user[1], user[2], user[3]]) # Discord Name, In-Game Name, Rank, LP
                else:
                    dataIndex = 0
                    for index, player in enumerate(midPlayers):
                        if player[0] < userNum:
                            dataIndex = index
                            break
                    midPlayers.insert(dataIndex, [userNum, user[0], user[1], user[2], user[3]])
            elif user[5] == 'ADC': 
                if len(adcPlayers) == 0:
                    adcPlayers.append([userNum, user[0], user[1], user[2], user[3]]) # Discord Name, In-Game Name, Rank, LP
                else:
                    dataIndex = 0
                    for index, player in enumerate(adcPlayers):
                        if player[0] < userNum:
                            dataIndex = index
                            break
                    adcPlayers.insert(dataIndex, [userNum, user[0], user[1], user[2], user[3]])
            elif user[5] == 'Support': 
                if len(supPlayers) == 0:
                    supPlayers.append([userNum, user[0], user[1], user[2], user[3]]) # Discord Name, In-Game Name, Rank, LP
                else:
                    dataIndex = 0
                    for index, player in enumerate(supPlayers):
                        if player[0] < userNum:
                            dataIndex = index
                            break
                    supPlayers.insert(dataIndex, [userNum, user[0], user[1], user[2], user[3]])
            elif user[5] == 'Fill': 
                if len(fillPlayers) == 0:
                    fillPlayers.append([userNum, user[0], user[1], user[2], user[3]]) # Discord Name, In-Game Name, Rank, LP
                else:
                    dataIndex = 0
                    for index, player in enumerate(fillPlayers):
                        if player[0] < userNum:
                            dataIndex = index
                            break
                    fillPlayers.insert(dataIndex, [userNum, user[0], user[1], user[2], user[3]])
            else: 
                print("Invalid Role '{role}' Specified! Check Input".format(role = user[5]))
        else:
            print("Unrecognized Rank '{rank}'. Check Data".format(rank = user[2]))

    # Setup Top, Mid, Jungle Headers
    ## Top
    primarySheet.cell(row=currentRow, column=currentCol).font = Font(name="Roboto Mono", bold=True)
    primarySheet.cell(row=currentRow, column=currentCol).value = "Primary Top Laners"
    primarySheet.cell(row=currentRow+1, column=currentCol).font = Font(name="Roboto Mono", underline="single")
    primarySheet.cell(row=currentRow+1, column=currentCol).value = "Discord Name"
    primarySheet.cell(row=currentRow+1, column=currentCol+1).font = Font(name="Roboto Mono", underline="single")
    primarySheet.cell(row=currentRow+1, column=currentCol+1).value = "In-Game Name"
    primarySheet.cell(row=currentRow+1, column=currentCol+2).font = Font(name="Roboto Mono", underline="single")
    primarySheet.cell(row=currentRow+1, column=currentCol+2).value = "Rank"
    primarySheet.cell(row=currentRow+1, column=currentCol+3).font = Font(name="Roboto Mono", underline="single")
    primarySheet.cell(row=currentRow+1, column=currentCol+3).value = "LP"

    ## Mid
    currentCol = 6
    primarySheet.cell(row=currentRow, column=currentCol).font = Font(name="Roboto Mono", bold=True)
    primarySheet.cell(row=currentRow, column=currentCol).value = "Primary Mid Laners"
    primarySheet.cell(row=currentRow+1, column=currentCol).font = Font(name="Roboto Mono", underline="single")
    primarySheet.cell(row=currentRow+1, column=currentCol).value = "Discord Name"
    primarySheet.cell(row=currentRow+1, column=currentCol+1).font = Font(name="Roboto Mono", underline="single")
    primarySheet.cell(row=currentRow+1, column=currentCol+1).value = "In-Game Name"
    primarySheet.cell(row=currentRow+1, column=currentCol+2).font = Font(name="Roboto Mono", underline="single")
    primarySheet.cell(row=currentRow+1, column=currentCol+2).value = "Rank"
    primarySheet.cell(row=currentRow+1, column=currentCol+3).font = Font(name="Roboto Mono", underline="single")
    primarySheet.cell(row=currentRow+1, column=currentCol+3).value = "LP"
    
    ## Jungle
    currentCol = 11
    primarySheet.cell(row=currentRow, column=currentCol).font = Font(name="Roboto Mono", bold=True)
    primarySheet.cell(row=currentRow, column=currentCol).value = "Primary Junglers"
    primarySheet.cell(row=currentRow+1, column=currentCol).font = Font(name="Roboto Mono", underline="single")
    primarySheet.cell(row=currentRow+1, column=currentCol).value = "Discord Name"
    primarySheet.cell(row=currentRow+1, column=currentCol+1).font = Font(name="Roboto Mono", underline="single")
    primarySheet.cell(row=currentRow+1, column=currentCol+1).value = "In-Game Name"
    primarySheet.cell(row=currentRow+1, column=currentCol+2).font = Font(name="Roboto Mono", underline="single")
    primarySheet.cell(row=currentRow+1, column=currentCol+2).value = "Rank"
    primarySheet.cell(row=currentRow+1, column=currentCol+3).font = Font(name="Roboto Mono", underline="single")
    primarySheet.cell(row=currentRow+1, column=currentCol+3).value = "LP"

    currentRow = 3
    currentCol = 1
    for player in topPlayers:
        if player[0] != 0:
            primarySheet.cell(row=currentRow, column=currentCol).font = Font(name="Roboto Mono")
            primarySheet.cell(row=currentRow, column=currentCol).value = player[1]
            primarySheet.cell(row=currentRow, column=currentCol+1).font = Font(name="Roboto Mono")
            primarySheet.cell(row=currentRow, column=currentCol+1).value = player[2]
            primarySheet.cell(row=currentRow, column=currentCol+2).font = Font(name="Roboto Mono")
            primarySheet.cell(row=currentRow, column=currentCol+2).value = player[3]
            primarySheet.cell(row=currentRow, column=currentCol+3).font = Font(name="Roboto Mono")
            primarySheet.cell(row=currentRow, column=currentCol+3).value = player[4]
        currentRow += 1
    adcRow = currentRow + 1

    currentRow = 3
    currentCol = 6
    for player in midPlayers:
        if player[0] != 0:
            primarySheet.cell(row=currentRow, column=currentCol).font = Font(name="Roboto Mono")
            primarySheet.cell(row=currentRow, column=currentCol).value = player[1]
            primarySheet.cell(row=currentRow, column=currentCol+1).font = Font(name="Roboto Mono")
            primarySheet.cell(row=currentRow, column=currentCol+1).value = player[2]
            primarySheet.cell(row=currentRow, column=currentCol+2).font = Font(name="Roboto Mono")
            primarySheet.cell(row=currentRow, column=currentCol+2).value = player[3]
            primarySheet.cell(row=currentRow, column=currentCol+3).font = Font(name="Roboto Mono")
            primarySheet.cell(row=currentRow, column=currentCol+3).value = player[4]
        currentRow += 1
    supRow = currentRow + 1

    currentRow = 3
    currentCol = 11
    for player in junglePlayers:
        if player[0] != 0:
            primarySheet.cell(row=currentRow, column=currentCol).font = Font(name="Roboto Mono")
            primarySheet.cell(row=currentRow, column=currentCol).value = player[1]
            primarySheet.cell(row=currentRow, column=currentCol+1).font = Font(name="Roboto Mono")
            primarySheet.cell(row=currentRow, column=currentCol+1).value = player[2]
            primarySheet.cell(row=currentRow, column=currentCol+2).font = Font(name="Roboto Mono")
            primarySheet.cell(row=currentRow, column=currentCol+2).value = player[3]
            primarySheet.cell(row=currentRow, column=currentCol+3).font = Font(name="Roboto Mono")
            primarySheet.cell(row=currentRow, column=currentCol+3).value = player[4]
        currentRow += 1
    fillRow = currentRow + 1

    ## ADC
    currentRow = adcRow
    currentCol = 1
    primarySheet.cell(row=currentRow, column=currentCol).font = Font(name="Roboto Mono", bold=True)
    primarySheet.cell(row=currentRow, column=currentCol).value = "Primary ADC Players"
    primarySheet.cell(row=currentRow+1, column=currentCol).font = Font(name="Roboto Mono", underline="single")
    primarySheet.cell(row=currentRow+1, column=currentCol).value = "Discord Name"
    primarySheet.cell(row=currentRow+1, column=currentCol+1).font = Font(name="Roboto Mono", underline="single")
    primarySheet.cell(row=currentRow+1, column=currentCol+1).value = "In-Game Name"
    primarySheet.cell(row=currentRow+1, column=currentCol+2).font = Font(name="Roboto Mono", underline="single")
    primarySheet.cell(row=currentRow+1, column=currentCol+2).value = "Rank"
    primarySheet.cell(row=currentRow+1, column=currentCol+3).font = Font(name="Roboto Mono", underline="single")
    primarySheet.cell(row=currentRow+1, column=currentCol+3).value = "LP"
    currentRow += 2

    for player in adcPlayers:
        if player[0] != 0:
            primarySheet.cell(row=currentRow, column=currentCol).font = Font(name="Roboto Mono")
            primarySheet.cell(row=currentRow, column=currentCol).value = player[1]
            primarySheet.cell(row=currentRow, column=currentCol+1).font = Font(name="Roboto Mono")
            primarySheet.cell(row=currentRow, column=currentCol+1).value = player[2]
            primarySheet.cell(row=currentRow, column=currentCol+2).font = Font(name="Roboto Mono")
            primarySheet.cell(row=currentRow, column=currentCol+2).value = player[3]
            primarySheet.cell(row=currentRow, column=currentCol+3).font = Font(name="Roboto Mono")
            primarySheet.cell(row=currentRow, column=currentCol+3).value = player[4]
        currentRow += 1

    ## Support
    currentRow = supRow
    currentCol = 6
    primarySheet.cell(row=currentRow, column=currentCol).font = Font(name="Roboto Mono", bold=True)
    primarySheet.cell(row=currentRow, column=currentCol).value = "Primary Support Players"
    primarySheet.cell(row=currentRow+1, column=currentCol).font = Font(name="Roboto Mono", underline="single")
    primarySheet.cell(row=currentRow+1, column=currentCol).value = "Discord Name"
    primarySheet.cell(row=currentRow+1, column=currentCol+1).font = Font(name="Roboto Mono", underline="single")
    primarySheet.cell(row=currentRow+1, column=currentCol+1).value = "In-Game Name"
    primarySheet.cell(row=currentRow+1, column=currentCol+2).font = Font(name="Roboto Mono", underline="single")
    primarySheet.cell(row=currentRow+1, column=currentCol+2).value = "Rank"
    primarySheet.cell(row=currentRow+1, column=currentCol+3).font = Font(name="Roboto Mono", underline="single")
    primarySheet.cell(row=currentRow+1, column=currentCol+3).value = "LP"
    currentRow += 2

    for player in supPlayers:
        if player[0] != 0:
            primarySheet.cell(row=currentRow, column=currentCol).font = Font(name="Roboto Mono")
            primarySheet.cell(row=currentRow, column=currentCol).value = player[1]
            primarySheet.cell(row=currentRow, column=currentCol+1).font = Font(name="Roboto Mono")
            primarySheet.cell(row=currentRow, column=currentCol+1).value = player[2]
            primarySheet.cell(row=currentRow, column=currentCol+2).font = Font(name="Roboto Mono")
            primarySheet.cell(row=currentRow, column=currentCol+2).value = player[3]
            primarySheet.cell(row=currentRow, column=currentCol+3).font = Font(name="Roboto Mono")
            primarySheet.cell(row=currentRow, column=currentCol+3).value = player[4]
        currentRow += 1

    ## Fill
    currentRow = fillRow
    currentCol = 11
    primarySheet.cell(row=currentRow, column=currentCol).font = Font(name="Roboto Mono", bold=True)
    primarySheet.cell(row=currentRow, column=currentCol).value = "Primary Fill Players"
    primarySheet.cell(row=currentRow+1, column=currentCol).font = Font(name="Roboto Mono", underline="single")
    primarySheet.cell(row=currentRow+1, column=currentCol).value = "Discord Name"
    primarySheet.cell(row=currentRow+1, column=currentCol+1).font = Font(name="Roboto Mono", underline="single")
    primarySheet.cell(row=currentRow+1, column=currentCol+1).value = "In-Game Name"
    primarySheet.cell(row=currentRow+1, column=currentCol+2).font = Font(name="Roboto Mono", underline="single")
    primarySheet.cell(row=currentRow+1, column=currentCol+2).value = "Rank"
    primarySheet.cell(row=currentRow+1, column=currentCol+3).font = Font(name="Roboto Mono", underline="single")
    primarySheet.cell(row=currentRow+1, column=currentCol+3).value = "LP"
    currentRow += 2

    for player in fillPlayers:
        if player[0] != 0:
            primarySheet.cell(row=currentRow, column=currentCol).font = Font(name="Roboto Mono")
            primarySheet.cell(row=currentRow, column=currentCol).value = player[1]
            primarySheet.cell(row=currentRow, column=currentCol+1).font = Font(name="Roboto Mono")
            primarySheet.cell(row=currentRow, column=currentCol+1).value = player[2]
            primarySheet.cell(row=currentRow, column=currentCol+2).font = Font(name="Roboto Mono")
            primarySheet.cell(row=currentRow, column=currentCol+2).value = player[3]
            primarySheet.cell(row=currentRow, column=currentCol+3).font = Font(name="Roboto Mono")
            primarySheet.cell(row=currentRow, column=currentCol+3).value = player[4]
        currentRow += 1

    # Modify Column Dimensions to Look More Visually Pleasing
    primarySheet.column_dimensions['A'].width = 25
    primarySheet.column_dimensions['B'].width = 25
    primarySheet.column_dimensions['C'].width = 15
    primarySheet.column_dimensions['D'].width = 7
    primarySheet.column_dimensions['F'].width = 25
    primarySheet.column_dimensions['G'].width = 25
    primarySheet.column_dimensions['H'].width = 15
    primarySheet.column_dimensions['I'].width = 7
    primarySheet.column_dimensions['K'].width = 25
    primarySheet.column_dimensions['L'].width = 25
    primarySheet.column_dimensions['M'].width = 15
    primarySheet.column_dimensions['N'].width = 7

    print("Successfully Generated Primary Role Sheet!")
    return

def CreateSecondaryRoleWorksheet(userData, workbook):
    # Desc: Creates a worksheet that lists all users by their selected Secondary Roles and then Ranks
    
    secondarySheet = workbook.create_sheet("Secondary Roles")
    currentRow = 1
    currentCol = 1
    adcRow = 6
    supRow = 6
    fillRow = 6
    topPlayers = [[0, 'None', 'None', 'None', 'None']]
    junglePlayers = [[0, 'None', 'None', 'None', 'None']]
    midPlayers = [[0, 'None', 'None', 'None', 'None']]
    adcPlayers = [[0, 'None', 'None', 'None', 'None']]
    supPlayers = [[0, 'None', 'None', 'None', 'None']]
    fillPlayers = [[0, 'None', 'None', 'None', 'None']]

    # Sort Valid Players into their respective ranks
    for user in userData:
        userNum = 0
        userNum = CONST_RANK_NUM[user[2]]
        if userNum != 0:
            if user[6] == 'Top Lane': 
                if len(topPlayers) == 0:
                    topPlayers.append([userNum, user[0], user[1], user[2], user[3]]) # Discord Name, In-Game Name, Rank, LP
                else:
                    dataIndex = 0
                    for index, player in enumerate(topPlayers):
                        if player[0] < userNum:
                            dataIndex = index
                            break
                    topPlayers.insert(dataIndex, [userNum, user[0], user[1], user[2], user[3]])
            elif user[6] == 'Jungle': 
                if len(junglePlayers) == 0:
                    junglePlayers.append([userNum, user[0], user[1], user[2], user[3]]) # Discord Name, In-Game Name, Rank, LP
                else:
                    dataIndex = 0
                    for index, player in enumerate(junglePlayers):
                        if player[0] < userNum:
                            dataIndex = index
                            break
                    junglePlayers.insert(dataIndex, [userNum, user[0], user[1], user[2], user[3]])
            elif user[6] == 'Mid Lane': 
                if len(midPlayers) == 0:
                    midPlayers.append([userNum, user[0], user[1], user[2], user[3]]) # Discord Name, In-Game Name, Rank, LP
                else:
                    dataIndex = 0
                    for index, player in enumerate(midPlayers):
                        if player[0] < userNum:
                            dataIndex = index
                            break
                    midPlayers.insert(dataIndex, [userNum, user[0], user[1], user[2], user[3]])
            elif user[6] == 'ADC': 
                if len(adcPlayers) == 0:
                    adcPlayers.append([userNum, user[0], user[1], user[2], user[3]]) # Discord Name, In-Game Name, Rank, LP
                else:
                    dataIndex = 0
                    for index, player in enumerate(adcPlayers):
                        if player[0] < userNum:
                            dataIndex = index
                            break
                    adcPlayers.insert(dataIndex, [userNum, user[0], user[1], user[2], user[3]])
            elif user[6] == 'Support': 
                if len(supPlayers) == 0:
                    supPlayers.append([userNum, user[0], user[1], user[2], user[3]]) # Discord Name, In-Game Name, Rank, LP
                else:
                    dataIndex = 0
                    for index, player in enumerate(supPlayers):
                        if player[0] < userNum:
                            dataIndex = index
                            break
                    supPlayers.insert(dataIndex, [userNum, user[0], user[1], user[2], user[3]])
            elif user[6] == 'Fill': 
                if len(fillPlayers) == 0:
                    fillPlayers.append([userNum, user[0], user[1], user[2], user[3]]) # Discord Name, In-Game Name, Rank, LP
                else:
                    dataIndex = 0
                    for index, player in enumerate(fillPlayers):
                        if player[0] < userNum:
                            dataIndex = index
                            break
                    fillPlayers.insert(dataIndex, [userNum, user[0], user[1], user[2], user[3]])
            else: 
                print("Invalid Role '{role}' Specified! Check Input".format(role = user[5]))
        else:
            print("Unrecognized Rank '{rank}'. Check Data".format(rank = user[2]))

    # Setup Top, Mid, Jungle Headers
    ## Top
    secondarySheet.cell(row=currentRow, column=currentCol).font = Font(name="Roboto Mono", bold=True)
    secondarySheet.cell(row=currentRow, column=currentCol).value = "Secondary Top Laners"
    secondarySheet.cell(row=currentRow+1, column=currentCol).font = Font(name="Roboto Mono", underline="single")
    secondarySheet.cell(row=currentRow+1, column=currentCol).value = "Discord Name"
    secondarySheet.cell(row=currentRow+1, column=currentCol+1).font = Font(name="Roboto Mono", underline="single")
    secondarySheet.cell(row=currentRow+1, column=currentCol+1).value = "In-Game Name"
    secondarySheet.cell(row=currentRow+1, column=currentCol+2).font = Font(name="Roboto Mono", underline="single")
    secondarySheet.cell(row=currentRow+1, column=currentCol+2).value = "Rank"
    secondarySheet.cell(row=currentRow+1, column=currentCol+3).font = Font(name="Roboto Mono", underline="single")
    secondarySheet.cell(row=currentRow+1, column=currentCol+3).value = "LP"

    ## Mid
    currentCol = 6
    secondarySheet.cell(row=currentRow, column=currentCol).font = Font(name="Roboto Mono", bold=True)
    secondarySheet.cell(row=currentRow, column=currentCol).value = "Secondary Mid Laners"
    secondarySheet.cell(row=currentRow+1, column=currentCol).font = Font(name="Roboto Mono", underline="single")
    secondarySheet.cell(row=currentRow+1, column=currentCol).value = "Discord Name"
    secondarySheet.cell(row=currentRow+1, column=currentCol+1).font = Font(name="Roboto Mono", underline="single")
    secondarySheet.cell(row=currentRow+1, column=currentCol+1).value = "In-Game Name"
    secondarySheet.cell(row=currentRow+1, column=currentCol+2).font = Font(name="Roboto Mono", underline="single")
    secondarySheet.cell(row=currentRow+1, column=currentCol+2).value = "Rank"
    secondarySheet.cell(row=currentRow+1, column=currentCol+3).font = Font(name="Roboto Mono", underline="single")
    secondarySheet.cell(row=currentRow+1, column=currentCol+3).value = "LP"
    
    ## Jungle
    currentCol = 11
    secondarySheet.cell(row=currentRow, column=currentCol).font = Font(name="Roboto Mono", bold=True)
    secondarySheet.cell(row=currentRow, column=currentCol).value = "Secondary Junglers"
    secondarySheet.cell(row=currentRow+1, column=currentCol).font = Font(name="Roboto Mono", underline="single")
    secondarySheet.cell(row=currentRow+1, column=currentCol).value = "Discord Name"
    secondarySheet.cell(row=currentRow+1, column=currentCol+1).font = Font(name="Roboto Mono", underline="single")
    secondarySheet.cell(row=currentRow+1, column=currentCol+1).value = "In-Game Name"
    secondarySheet.cell(row=currentRow+1, column=currentCol+2).font = Font(name="Roboto Mono", underline="single")
    secondarySheet.cell(row=currentRow+1, column=currentCol+2).value = "Rank"
    secondarySheet.cell(row=currentRow+1, column=currentCol+3).font = Font(name="Roboto Mono", underline="single")
    secondarySheet.cell(row=currentRow+1, column=currentCol+3).value = "LP"

    currentRow = 3
    currentCol = 1
    for player in topPlayers:
        if player[0] != 0:
            secondarySheet.cell(row=currentRow, column=currentCol).font = Font(name="Roboto Mono")
            secondarySheet.cell(row=currentRow, column=currentCol).value = player[1]
            secondarySheet.cell(row=currentRow, column=currentCol+1).font = Font(name="Roboto Mono")
            secondarySheet.cell(row=currentRow, column=currentCol+1).value = player[2]
            secondarySheet.cell(row=currentRow, column=currentCol+2).font = Font(name="Roboto Mono")
            secondarySheet.cell(row=currentRow, column=currentCol+2).value = player[3]
            secondarySheet.cell(row=currentRow, column=currentCol+3).font = Font(name="Roboto Mono")
            secondarySheet.cell(row=currentRow, column=currentCol+3).value = player[4]
        currentRow += 1
    adcRow = currentRow + 1

    currentRow = 3
    currentCol = 6
    for player in midPlayers:
        if player[0] != 0:
            secondarySheet.cell(row=currentRow, column=currentCol).font = Font(name="Roboto Mono")
            secondarySheet.cell(row=currentRow, column=currentCol).value = player[1]
            secondarySheet.cell(row=currentRow, column=currentCol+1).font = Font(name="Roboto Mono")
            secondarySheet.cell(row=currentRow, column=currentCol+1).value = player[2]
            secondarySheet.cell(row=currentRow, column=currentCol+2).font = Font(name="Roboto Mono")
            secondarySheet.cell(row=currentRow, column=currentCol+2).value = player[3]
            secondarySheet.cell(row=currentRow, column=currentCol+3).font = Font(name="Roboto Mono")
            secondarySheet.cell(row=currentRow, column=currentCol+3).value = player[4]
        currentRow += 1
    supRow = currentRow + 1

    currentRow = 3
    currentCol = 11
    for player in junglePlayers:
        if player[0] != 0:
            secondarySheet.cell(row=currentRow, column=currentCol).font = Font(name="Roboto Mono")
            secondarySheet.cell(row=currentRow, column=currentCol).value = player[1]
            secondarySheet.cell(row=currentRow, column=currentCol+1).font = Font(name="Roboto Mono")
            secondarySheet.cell(row=currentRow, column=currentCol+1).value = player[2]
            secondarySheet.cell(row=currentRow, column=currentCol+2).font = Font(name="Roboto Mono")
            secondarySheet.cell(row=currentRow, column=currentCol+2).value = player[3]
            secondarySheet.cell(row=currentRow, column=currentCol+3).font = Font(name="Roboto Mono")
            secondarySheet.cell(row=currentRow, column=currentCol+3).value = player[4]
        currentRow += 1
    fillRow = currentRow + 1

    ## ADC
    currentRow = adcRow
    currentCol = 1
    secondarySheet.cell(row=currentRow, column=currentCol).font = Font(name="Roboto Mono", bold=True)
    secondarySheet.cell(row=currentRow, column=currentCol).value = "Secondary ADC Players"
    secondarySheet.cell(row=currentRow+1, column=currentCol).font = Font(name="Roboto Mono", underline="single")
    secondarySheet.cell(row=currentRow+1, column=currentCol).value = "Discord Name"
    secondarySheet.cell(row=currentRow+1, column=currentCol+1).font = Font(name="Roboto Mono", underline="single")
    secondarySheet.cell(row=currentRow+1, column=currentCol+1).value = "In-Game Name"
    secondarySheet.cell(row=currentRow+1, column=currentCol+2).font = Font(name="Roboto Mono", underline="single")
    secondarySheet.cell(row=currentRow+1, column=currentCol+2).value = "Rank"
    secondarySheet.cell(row=currentRow+1, column=currentCol+3).font = Font(name="Roboto Mono", underline="single")
    secondarySheet.cell(row=currentRow+1, column=currentCol+3).value = "LP"
    currentRow += 2

    for player in adcPlayers:
        if player[0] != 0:
            secondarySheet.cell(row=currentRow, column=currentCol).font = Font(name="Roboto Mono")
            secondarySheet.cell(row=currentRow, column=currentCol).value = player[1]
            secondarySheet.cell(row=currentRow, column=currentCol+1).font = Font(name="Roboto Mono")
            secondarySheet.cell(row=currentRow, column=currentCol+1).value = player[2]
            secondarySheet.cell(row=currentRow, column=currentCol+2).font = Font(name="Roboto Mono")
            secondarySheet.cell(row=currentRow, column=currentCol+2).value = player[3]
            secondarySheet.cell(row=currentRow, column=currentCol+3).font = Font(name="Roboto Mono")
            secondarySheet.cell(row=currentRow, column=currentCol+3).value = player[4]
        currentRow += 1

    ## Support
    currentRow = supRow
    currentCol = 6
    secondarySheet.cell(row=currentRow, column=currentCol).font = Font(name="Roboto Mono", bold=True)
    secondarySheet.cell(row=currentRow, column=currentCol).value = "Secondary Support Players"
    secondarySheet.cell(row=currentRow+1, column=currentCol).font = Font(name="Roboto Mono", underline="single")
    secondarySheet.cell(row=currentRow+1, column=currentCol).value = "Discord Name"
    secondarySheet.cell(row=currentRow+1, column=currentCol+1).font = Font(name="Roboto Mono", underline="single")
    secondarySheet.cell(row=currentRow+1, column=currentCol+1).value = "In-Game Name"
    secondarySheet.cell(row=currentRow+1, column=currentCol+2).font = Font(name="Roboto Mono", underline="single")
    secondarySheet.cell(row=currentRow+1, column=currentCol+2).value = "Rank"
    secondarySheet.cell(row=currentRow+1, column=currentCol+3).font = Font(name="Roboto Mono", underline="single")
    secondarySheet.cell(row=currentRow+1, column=currentCol+3).value = "LP"
    currentRow += 2

    for player in supPlayers:
        if player[0] != 0:
            secondarySheet.cell(row=currentRow, column=currentCol).font = Font(name="Roboto Mono")
            secondarySheet.cell(row=currentRow, column=currentCol).value = player[1]
            secondarySheet.cell(row=currentRow, column=currentCol+1).font = Font(name="Roboto Mono")
            secondarySheet.cell(row=currentRow, column=currentCol+1).value = player[2]
            secondarySheet.cell(row=currentRow, column=currentCol+2).font = Font(name="Roboto Mono")
            secondarySheet.cell(row=currentRow, column=currentCol+2).value = player[3]
            secondarySheet.cell(row=currentRow, column=currentCol+3).font = Font(name="Roboto Mono")
            secondarySheet.cell(row=currentRow, column=currentCol+3).value = player[4]
        currentRow += 1

    ## Fill
    currentRow = fillRow
    currentCol = 11
    secondarySheet.cell(row=currentRow, column=currentCol).font = Font(name="Roboto Mono", bold=True)
    secondarySheet.cell(row=currentRow, column=currentCol).value = "Secondary Fill Players"
    secondarySheet.cell(row=currentRow+1, column=currentCol).font = Font(name="Roboto Mono", underline="single")
    secondarySheet.cell(row=currentRow+1, column=currentCol).value = "Discord Name"
    secondarySheet.cell(row=currentRow+1, column=currentCol+1).font = Font(name="Roboto Mono", underline="single")
    secondarySheet.cell(row=currentRow+1, column=currentCol+1).value = "In-Game Name"
    secondarySheet.cell(row=currentRow+1, column=currentCol+2).font = Font(name="Roboto Mono", underline="single")
    secondarySheet.cell(row=currentRow+1, column=currentCol+2).value = "Rank"
    secondarySheet.cell(row=currentRow+1, column=currentCol+3).font = Font(name="Roboto Mono", underline="single")
    secondarySheet.cell(row=currentRow+1, column=currentCol+3).value = "LP"
    currentRow += 2

    for player in fillPlayers:
        if player[0] != 0:
            secondarySheet.cell(row=currentRow, column=currentCol).font = Font(name="Roboto Mono")
            secondarySheet.cell(row=currentRow, column=currentCol).value = player[1]
            secondarySheet.cell(row=currentRow, column=currentCol+1).font = Font(name="Roboto Mono")
            secondarySheet.cell(row=currentRow, column=currentCol+1).value = player[2]
            secondarySheet.cell(row=currentRow, column=currentCol+2).font = Font(name="Roboto Mono")
            secondarySheet.cell(row=currentRow, column=currentCol+2).value = player[3]
            secondarySheet.cell(row=currentRow, column=currentCol+3).font = Font(name="Roboto Mono")
            secondarySheet.cell(row=currentRow, column=currentCol+3).value = player[4]
        currentRow += 1

    # Modify Column Dimensions to Look More Visually Pleasing
    secondarySheet.column_dimensions['A'].width = 25
    secondarySheet.column_dimensions['B'].width = 25
    secondarySheet.column_dimensions['C'].width = 15
    secondarySheet.column_dimensions['D'].width = 7
    secondarySheet.column_dimensions['F'].width = 25
    secondarySheet.column_dimensions['G'].width = 25
    secondarySheet.column_dimensions['H'].width = 15
    secondarySheet.column_dimensions['I'].width = 7
    secondarySheet.column_dimensions['K'].width = 25
    secondarySheet.column_dimensions['L'].width = 25
    secondarySheet.column_dimensions['M'].width = 15
    secondarySheet.column_dimensions['N'].width = 7

    print("Successfully Generated Secondary Role Sheet!")
    return

def CreatePlayerDatabaseWorksheet(userData, workbook):
    # Desc: Creates a worksheet that lists all users and their info by their corresponding ranks in descending order
    
    databaseSheet = workbook.create_sheet("User Reference")
    currentRow = 1
    currentCol = 1
    
    # Sort Valid Users by their Rank and then display their Discord Username, In-Game Name and Rank
    sortedUsers = [[0, 'None', 'None', 'None', 'None']]

    for user in userData:
        userNum = 0
        userNum = CONST_RANK_NUM[user[2]]
        if userNum != 0:
            if len(sortedUsers) == 0:
                sortedUsers.append([userNum, user[0], user[1], user[2], user[3]]) # Discord Name, In-Game Name, Rank, LP
            else:
                dataIndex = 0
                for index, sortedUser in enumerate(sortedUsers):
                    if sortedUser[0] < userNum:
                        dataIndex = index
                        break
                sortedUsers.insert(dataIndex, [userNum, user[0], user[1], user[2], user[3]])
        else:
            print("Unrecognized Rank '{rank}'. Check Data".format(rank = user[2]))
    
    databaseSheet.cell(row=currentRow, column=currentCol).font = Font(name="Roboto Mono", bold=True, underline="single")
    databaseSheet.cell(row=currentRow, column=currentCol).value = "Discord Name"
    databaseSheet.cell(row=currentRow, column=currentCol+1).font = Font(name="Roboto Mono", bold=True, underline="single")
    databaseSheet.cell(row=currentRow, column=currentCol+1).value = "In-Game Name"
    databaseSheet.cell(row=currentRow, column=currentCol+2).font = Font(name="Roboto Mono", bold=True, underline="single")
    databaseSheet.cell(row=currentRow, column=currentCol+2).value = "Rank"
    databaseSheet.cell(row=currentRow, column=currentCol+3).font = Font(name="Roboto Mono", bold=True, underline="single")
    databaseSheet.cell(row=currentRow, column=currentCol+3).value = "LP"

    currentRow = 2
    for playerData in sortedUsers:
        if playerData[0] != 0:
            databaseSheet.cell(row=currentRow, column=currentCol).font = Font(name="Roboto Mono")
            databaseSheet.cell(row=currentRow, column=currentCol).value = playerData[1]
            databaseSheet.cell(row=currentRow, column=currentCol+1).font = Font(name="Roboto Mono")
            databaseSheet.cell(row=currentRow, column=currentCol+1).value = playerData[2]
            databaseSheet.cell(row=currentRow, column=currentCol+2).font = Font(name="Roboto Mono")
            databaseSheet.cell(row=currentRow, column=currentCol+2).value = playerData[3]
            databaseSheet.cell(row=currentRow, column=currentCol+3).font = Font(name="Roboto Mono")
            databaseSheet.cell(row=currentRow, column=currentCol+3).value = playerData[4]
        currentRow += 1
    
    databaseSheet.column_dimensions['A'].width = 25
    databaseSheet.column_dimensions['B'].width = 25
    databaseSheet.column_dimensions['C'].width = 15
    databaseSheet.column_dimensions['D'].width = 7

    print("Successfully Generated User Reference Sheet!")
    return

def ProcessWorkbook(workbook):
    # Desc: Handles all of the workbook processing functions and updating
    worksheet = workbook.active
    rowMax = 1
    while(worksheet.cell(row=rowMax, column=1).value != None):
        rowMax += 1
    validData = GetValidEntries(worksheet, rowMax)
    CreateTeamWorksheet(validData, workbook)
    CreatePrimaryRoleWorksheet(validData, workbook)
    CreateSecondaryRoleWorksheet(validData, workbook)
    CreatePlayerDatabaseWorksheet(validData, workbook)
    return workbook

def SaveWorkbook(workbook, fileName):
    # Desc: Saves the sorted workbook with its modifications to the SortedForms Folder
    savePath = 'SortedForms/(Sorted)' + fileName
    workbook.save(savePath)

    print("Successfully Exported Workbook to Sorted Excel File!\n")
    return

if __name__ == "__main__":
    print("\nProgram Start")
    excelFiles = GetFileList()
    print("\n----Valid ExcelFiles List----")
    for fileInfo in excelFiles:
        print(fileInfo)
        workbookInfo = LoadWorkbook(fileInfo[0], fileInfo[1]) # Returns <Path to Input XLSX> <XLSX File Name> <Workbook Object>
        processedWorkbook = ProcessWorkbook(workbookInfo[2])
        SaveWorkbook(processedWorkbook, fileInfo[1])
    print("Program End")
